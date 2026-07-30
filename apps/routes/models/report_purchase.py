from datetime import datetime

from sqlalchemy import func
from apps.database.db_workshops import Workshops

from apps.database.db_products import Products
from apps.database.db_purchases import Purchases
from apps.database.db_suppliers import Suppliers
from apps.database.db_purchase_details import PurchaseDetails

from apps.utilities.responseHelpers import *
from apps.utilities.validators import role_validator
from apps.utilities.formatter import format_date

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from flask import send_file
from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
      SimpleDocTemplate,
      Table,
      TableStyle,
      Paragraph,
      Spacer
)

# REPORT SALES HELPER ============================================================ Begin
def _report_purchase_helper(workshop_id, supplier_id="", start_date="", end_date="" ):

      purchases = Purchases.query.filter(
            Purchases.workshop_id == workshop_id,
            Purchases.is_delete == 0
      )

      # Filter Supplier
      if supplier_id != "":
            purchases = purchases.filter(
                  Purchases.supplier_id == supplier_id
            )

      # Filter Tanggal
      if start_date and end_date:
            purchases = purchases.filter(
                  Purchases.purchase_date >= start_date,
                  Purchases.purchase_date <= end_date
            )

      return purchases
# REPORT SALES HELPER ============================================================ End

# FILTER DATE HELPER ============================================================ Begin
def _get_filter_date(start_date="", end_date=""):

      today = datetime.now()

      # Default hari ini
      if not start_date or not end_date:

            start_datetime = datetime(
                  today.year,
                  today.month,
                  today.day,
                  0,
                  0,
                  0
            )

            end_datetime = datetime(
                  today.year,
                  today.month,
                  today.day,
                  23,
                  59,
                  59
            )

      # Berdasarkan filter user
      else:

            start_datetime = datetime.strptime(
                  start_date,
                  "%Y-%m-%d"
            )

            end_datetime = datetime.strptime(
                  end_date,
                  "%Y-%m-%d"
            ).replace(
                  hour=23,
                  minute=59,
                  second=59
            )

      return (
            int(start_datetime.timestamp() * 1000),
            int(end_datetime.timestamp() * 1000)
      )
# FILTER DATE HELPER ============================================================ End

# REPORT SALES ============================================================ Begin
class ReportPurchaseModels():
      # REPORT SUMMARY ============================================================ Begin
      def report_summary(user_role,user_id,workshop_id,supplier_id="",start_date=None,end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Data ---------------------------------------- Start
                  purchases = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  ).all()
                  # Get Data ---------------------------------------- Finish
                  # Summary ---------------------------------------- Start
                  total_transaction = len(purchases)

                  total_purchase = sum(
                        purchase.total
                        for purchase in purchases
                  )

                  total_item = sum(
                        detail.quantity
                        for purchase in purchases
                        for detail in purchase.purchase_details
                  )

                  active_supplier = len(
                        set(
                              purchase.supplier_id
                              for purchase in purchases
                              if purchase.supplier_id
                        )
                  )
                  # Summary ---------------------------------------- Finish

                  # Response ---------------------------------------- Start
                  return success_data(
                        data={
                              "total_transaction": total_transaction,
                              "total_purchase": total_purchase,
                              "active_supplier": active_supplier,
                              "total_item": total_item
                        },
                        status_code=200
                  )
                  # Response ---------------------------------------- Finish
            except Exception as e:
                  print("REPORT PURCHASE SUMMARY ERROR :", e)
                  raise
                  # return bad_request(str(e))
      # REPORT SUMMARY ============================================================ End

      # PURCHASE CHART ============================================================ Begin
      def purchase_chart(user_role, user_id, workshop_id, supplier_id="", start_date=None, end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Initialize Query ---------------------------------------- Start
                  query = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  )
                  # Initialize Query ---------------------------------------- Finish

                  # Chart Query ---------------------------------------- Start
                  purchases = query.with_entities(
                        Purchases.purchase_date,
                        func.sum(Purchases.total).label("total_purchase")
                  ).group_by(
                        Purchases.purchase_date
                  ).order_by(
                        Purchases.purchase_date.asc()
                  ).all()

                  chart = []

                  for purchase in purchases:

                        chart.append({
                              "date": format_date(
                                    purchase.purchase_date
                              ),
                              "total_purchase": purchase.total_purchase
                        })
                  # Chart Query ---------------------------------------- Finish
                  
                  # Response ---------------------------------------- Start
                  return success_data(
                        data={
                              "chart": chart
                        },
                        status_code=200
                  )
                  # Response ---------------------------------------- Finish         
            except Exception as e:
                  return bad_request(str(e))
      # PURCHASE CHART ============================================================ End

      # TOP SUPPLIER ============================================================ Begin
      def top_suppliers(user_role,user_id,workshop_id,supplier_id="",start_date=None,end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Data ---------------------------------------- Start
                  query = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  )
                  # Get Data ---------------------------------------- Finish

                  # Top Suppliers ---------------------------------------- Start
                  suppliers = query.join(
                        Suppliers,
                        Purchases.supplier_id == Suppliers.id
                  ).with_entities(
                        Suppliers.name,
                        func.sum(Purchases.total).label(
                              "total_purchase"
                        )
                  ).group_by(
                        Suppliers.id,
                        Suppliers.name
                  ).order_by(
                        func.sum(Purchases.total).desc()
                  ).limit(5).all()

                  data = []

                  for supplier in suppliers:

                        data.append({
                              "name": supplier.name,
                              "total_purchase": supplier.total_purchase
                        })

                  # Top Suppliers ---------------------------------------- Finish
                  
                  # Response ---------------------------------------- Start
                  return success_data(
                        data={
                              "top_suppliers": data
                        },
                        status_code=200
                  )
                  # Response ---------------------------------------- Finish

            except Exception as e:
                  return bad_request(str(e))
      # TOP SUPPLIER ============================================================ End

      # TOP PRODUCT ============================================================ Begin
      def top_product(user_role, user_id, workshop_id, supplier_id="", start_date=None, end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Data ---------------------------------------- Start
                  query = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  )
                  # Get Data ---------------------------------------- Finish

                  # Top Products ---------------------------------------- Start
                  products = query.join(
                        PurchaseDetails,
                        Purchases.id == PurchaseDetails.purchase_id
                  ).join(
                        Products,
                        PurchaseDetails.product_id == Products.id
                  ).with_entities(
                        Products.product_name,
                        func.sum(
                              PurchaseDetails.quantity
                        ).label("total_quantity"),
                        func.sum(
                              PurchaseDetails.quantity *
                              PurchaseDetails.unit_cost
                        ).label("total_purchase")
                  ).group_by(
                        Products.id,
                        Products.product_name
                  ).order_by(
                        func.sum(
                              PurchaseDetails.quantity
                        ).desc()
                  ).limit(5).all()

                  data = []

                  for product in products:
                        data.append({
                              "product_name": product.product_name,
                              "total_quantity": product.total_quantity,
                              "total_purchase": product.total_purchase
                        })
                  # Top Products ---------------------------------------- Finish

                  # Response ---------------------------------------- Start
                  return success_data(
                        data={
                              "top_products": data
                        },
                        status_code=200
                  )
                  # Response ---------------------------------------- Finish

            except Exception as e:
                  return bad_request(str(e))
      # TOP PRODUCT ============================================================ End
      
      # REPORT TABLE ============================================================ Begin
      def report_table(user_role,user_id,workshop_id,supplier_id="",start_date=None,end_date=None):
            try:

                 # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Data ---------------------------------------- Start
                  purchases = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  ).order_by(
                        Purchases.purchase_date.desc()
                  ).all()

                  report = []

                  for purchase in purchases:

                        total_item = sum(
                              detail.quantity
                              for detail in purchase.purchase_details
                        )

                        report.append({
                              "id": purchase.id,
                              "invoice": f"PB-{purchase.id:06d}",
                              "purchase_date": format_date(
                                    purchase.purchase_date
                              ),
                              "name": (
                                    purchase.suppliers.name
                                    if purchase.suppliers
                                    else "-"
                              ),
                              "total_item": total_item,
                              "total": purchase.total
                        })

                  # Get Data ---------------------------------------- Finish

                  # Report ---------------------------------------- Start
                  
                  # Report ---------------------------------------- Finish
                  
                  # Response ---------------------------------------- Finish
                  return success_data(
                        data={
                              "report": report
                        },
                        status_code=200
                  )
            except Exception as e:
                  print("REPORT PURCHASE TABLE ERROR :", e)
                  return bad_request(str(e))
      # REPOT TABLE ============================================================ End

      # EXPORT EXCEL ============================================================ Begin
      def export_excel(user_role, user_id, workshop_id, supplier_id="", start_date=None, end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()

                  # Hanya owner
                  if str(user_role) != "1":
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Purchase Data ---------------------------------------- Start
                  purchases = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  ).order_by(
                        Purchases.purchase_date.desc()
                  ).all()
                  # Get Purchase Data ---------------------------------------- Finish

                  # Summary ---------------------------------------- Start
                  total_transaction = len(purchases)

                  total_item = sum(
                        detail.quantity
                        for purchase in purchases
                        for detail in purchase.purchase_details
                  )

                  total_purchase = sum(
                        purchase.total
                        for purchase in purchases
                  )
                  # Summary ---------------------------------------- Finish

                  # Create Workbook ---------------------------------------- Start
                  workbook = Workbook()
                  worksheet = workbook.active
                  worksheet.title = "Laporan Pembelian"

                  worksheet.merge_cells("A1:F1")

                  worksheet["A1"] = "LAPORAN PEMBELIAN"
                  worksheet["A1"].font = Font(
                        bold=True,
                        size=16
                  )
                  worksheet["A1"].alignment = Alignment(
                        horizontal="center"
                  )

                  worksheet.append([])

                  worksheet.append([
                        "Nama Bengkel",
                        workshop.workshop_name
                  ])

                  worksheet.append([
                        "Periode",
                        (
                              f"{format_date(start_date)} "
                              f"s.d. {format_date(end_date)}"
                        )
                  ])

                  worksheet.append([])

                  worksheet.append([
                        "No",
                        "Invoice",
                        "Tanggal",
                        "Supplier",
                        "Jumlah Item",
                        "Total Pembelian"
                  ])

                  # Header Style
                  for cell in worksheet[6]:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(
                              horizontal="center",
                              vertical="center"
                        )
                  # Create Workbook ---------------------------------------- Finish

                  # Fill Data ---------------------------------------- Start
                  for index, purchase in enumerate(
                        purchases,
                        start=1
                  ):
                        purchase_total_item = sum(
                              detail.quantity
                              for detail in purchase.purchase_details
                        )

                        worksheet.append([
                              index,
                              f"PB-{purchase.id:06d}",
                              format_date(
                                    purchase.purchase_date
                              ),
                              (
                                    purchase.suppliers.name
                                    if purchase.suppliers
                                    else "-"
                              ),
                              purchase_total_item,
                              (
                                    f"Rp "
                                    f"{int(purchase.total or 0):,}"
                              ).replace(",", ".")
                        ])
                  # Fill Data ---------------------------------------- Finish

                  # Summary Data ---------------------------------------- Start
                  worksheet.append([])

                  worksheet.append([
                        "Jumlah Transaksi",
                        total_transaction
                  ])

                  worksheet.append([
                        "Total Item",
                        total_item
                  ])

                  worksheet.append([
                        "Total Pembelian",
                        (
                              f"Rp "
                              f"{int(total_purchase or 0):,}"
                        ).replace(",", ".")
                  ])
                  # Summary Data ---------------------------------------- Finish

                  # Worksheet Configuration ---------------------------------------- Start
                  worksheet.column_dimensions["A"].width = 8
                  worksheet.column_dimensions["B"].width = 25
                  worksheet.column_dimensions["C"].width = 16
                  worksheet.column_dimensions["D"].width = 30
                  worksheet.column_dimensions["E"].width = 15
                  worksheet.column_dimensions["F"].width = 22

                  worksheet.freeze_panes = "A7"

                  for row in worksheet.iter_rows(
                        min_row=7,
                        max_row=worksheet.max_row,
                        min_col=1,
                        max_col=6
                  ):
                        row[0].alignment = Alignment(
                              horizontal="center"
                        )

                        row[2].alignment = Alignment(
                              horizontal="center"
                        )

                        row[4].alignment = Alignment(
                              horizontal="center"
                        )

                        row[5].alignment = Alignment(
                              horizontal="right"
                        )
                  # Worksheet Configuration ---------------------------------------- Finish

                  # Response File ---------------------------------------- Start
                  buffer = BytesIO()

                  workbook.save(buffer)
                  buffer.seek(0)

                  return send_file(
                        buffer,
                        as_attachment=True,
                        download_name="report_purchase.xlsx",
                        mimetype=(
                              "application/vnd.openxmlformats-"
                              "officedocument.spreadsheetml.sheet"
                        )
                  )
                  # Response File ---------------------------------------- Finish

            except Exception as e:
                  return bad_request(str(e))
      # EXPORT EXCEL ============================================================ End
      
      # EXPORT PDF ============================================================ Begin
      def export_pdf(user_role, user_id, workshop_id, supplier_id="", start_date=None, end_date=None):
            try:

                  # Access Validation ---------------------------------------- Start
                  access = role_validator(user_role)

                  if not access:
                        return authorization_error()

                  # Hanya owner
                  if str(user_role) != "1":
                        return authorization_error()
                  # Access Validation ---------------------------------------- Finish

                  # Check Workshop ---------------------------------------- Start
                  workshop = Workshops.query.filter_by(
                        id=workshop_id,
                        is_delete=0
                  ).first()

                  if not workshop:
                        return not_found(
                              "Workshop could not be found."
                        )
                  # Check Workshop ---------------------------------------- Finish

                  # Filter Date ---------------------------------------- Start
                  start_date, end_date = _get_filter_date(
                        start_date,
                        end_date
                  )
                  # Filter Date ---------------------------------------- Finish

                  # Get Purchase Data ---------------------------------------- Start
                  purchases = _report_purchase_helper(
                        workshop_id,
                        supplier_id,
                        start_date,
                        end_date
                  ).order_by(
                        Purchases.purchase_date.desc()
                  ).all()
                  # Get Purchase Data ---------------------------------------- Finish

                  # Summary ---------------------------------------- Start
                  total_transaction = len(purchases)

                  total_item = sum(
                        detail.quantity
                        for purchase in purchases
                        for detail in purchase.purchase_details
                  )

                  total_purchase = sum(
                        purchase.total
                        for purchase in purchases
                  )
                  # Summary ---------------------------------------- Finish

                  # Table Data ---------------------------------------- Start
                  table_data = [
                        [
                              "No",
                              "Invoice",
                              "Tanggal",
                              "Supplier",
                              "Jumlah Item",
                              "Total"
                        ]
                  ]

                  for index, purchase in enumerate(
                        purchases,
                        start=1
                  ):
                        purchase_total_item = sum(
                              detail.quantity
                              for detail in purchase.purchase_details
                        )

                        table_data.append([
                              index,
                              f"PB-{purchase.id:06d}",
                              format_date(
                                    purchase.purchase_date
                              ),
                              (
                                    purchase.suppliers.name
                                    if purchase.suppliers
                                    else "-"
                              ),
                              purchase_total_item,
                              (
                                    f"Rp "
                                    f"{int(purchase.total or 0):,}"
                              ).replace(",", ".")
                        ])

                  if len(purchases) == 0:
                        table_data.append([
                              "",
                              "Tidak ada data pembelian",
                              "",
                              "",
                              "",
                              ""
                        ])
                  # Table Data ---------------------------------------- Finish

                  # Initialize PDF ---------------------------------------- Start
                  buffer = BytesIO()

                  document = SimpleDocTemplate(
                        buffer,
                        pagesize=A4,
                        leftMargin=1.5 * cm,
                        rightMargin=1.5 * cm,
                        topMargin=1.5 * cm,
                        bottomMargin=1.5 * cm
                  )

                  styles = getSampleStyleSheet()
                  elements = []
                  # Initialize PDF ---------------------------------------- Finish

                  # PDF Header ---------------------------------------- Start
                  elements.append(
                        Paragraph(
                              "<b>LAPORAN PEMBELIAN</b>",
                              styles["Title"]
                        )
                  )

                  elements.append(
                        Spacer(1, 0.2 * cm)
                  )

                  elements.append(
                        Paragraph(
                              (
                                    f"<b>Nama Bengkel:</b> "
                                    f"{workshop.workshop_name}"
                              ),
                              styles["Normal"]
                        )
                  )

                  elements.append(
                        Paragraph(
                              (
                                    f"<b>Periode Laporan:</b> "
                                    f"{format_date(start_date)} "
                                    f"s.d. {format_date(end_date)}"
                              ),
                              styles["Normal"]
                        )
                  )

                  elements.append(
                        Spacer(1, 0.5 * cm)
                  )
                  # PDF Header ---------------------------------------- Finish

                  # PDF Table ---------------------------------------- Start
                  table = Table(
                        table_data,
                        colWidths=[
                              1 * cm,       # No
                              3.8 * cm,     # Invoice
                              2.6 * cm,     # Tanggal
                              4.2 * cm,     # Supplier
                              2.4 * cm,     # Jumlah Item
                              4 * cm        # Total
                        ],
                        repeatRows=1
                  )

                  table.setStyle(
                        TableStyle([
                              (
                                    "GRID",
                                    (0, 0),
                                    (-1, -1),
                                    0.5,
                                    colors.black
                              ),
                              (
                                    "BACKGROUND",
                                    (0, 0),
                                    (-1, 0),
                                    colors.HexColor("#0d6efd")
                              ),
                              (
                                    "TEXTCOLOR",
                                    (0, 0),
                                    (-1, 0),
                                    colors.white
                              ),
                              (
                                    "FONTNAME",
                                    (0, 0),
                                    (-1, 0),
                                    "Helvetica-Bold"
                              ),
                              (
                                    "FONTSIZE",
                                    (0, 0),
                                    (-1, -1),
                                    9
                              ),
                              (
                                    "ALIGN",
                                    (0, 0),
                                    (0, -1),
                                    "CENTER"
                              ),
                              (
                                    "ALIGN",
                                    (2, 0),
                                    (2, -1),
                                    "CENTER"
                              ),
                              (
                                    "ALIGN",
                                    (4, 0),
                                    (4, -1),
                                    "CENTER"
                              ),
                              (
                                    "ALIGN",
                                    (5, 1),
                                    (5, -1),
                                    "RIGHT"
                              ),
                              (
                                    "VALIGN",
                                    (0, 0),
                                    (-1, -1),
                                    "MIDDLE"
                              ),
                              (
                                    "TOPPADDING",
                                    (0, 0),
                                    (-1, -1),
                                    6
                              ),
                              (
                                    "BOTTOMPADDING",
                                    (0, 0),
                                    (-1, -1),
                                    6
                              )
                        ])
                  )

                  elements.append(table)
                  # PDF Table ---------------------------------------- Finish

                  # PDF Summary ---------------------------------------- Start
                  elements.append(
                        Spacer(1, 0.5 * cm)
                  )

                  elements.append(
                        Paragraph(
                              (
                                    f"<b>Jumlah Transaksi:</b> "
                                    f"{total_transaction}"
                              ),
                              styles["Normal"]
                        )
                  )

                  elements.append(
                        Paragraph(
                              (
                                    f"<b>Total Item:</b> "
                                    f"{total_item}"
                              ),
                              styles["Normal"]
                        )
                  )

                  elements.append(
                        Paragraph(
                              (
                                    f"<b>Total Pembelian:</b> "
                                    f"Rp "
                                    f"{int(total_purchase or 0):,}"
                              ).replace(",", "."),
                              styles["Normal"]
                        )
                  )
                  # PDF Summary ---------------------------------------- Finish

                  # Generate PDF ---------------------------------------- Start
                  document.build(elements)

                  buffer.seek(0)
                  # Generate PDF ---------------------------------------- Finish

                  # Return File ---------------------------------------- Start
                  return send_file(
                        buffer,
                        as_attachment=True,
                        download_name="report_purchase.pdf",
                        mimetype="application/pdf"
                  )
                  # Return File ---------------------------------------- Finish

            except Exception as e:
                  return bad_request(str(e))
      # EXPORT PDF ============================================================ End

# REPORT SALES ============================================================ End